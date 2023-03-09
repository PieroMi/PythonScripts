import pandas as pd
from pandas import io
from pandas.io import excel
from pandas.io.formats import style

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from openpyxl import worksheet
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.styles import numbers

workbook = openpyxl.load_workbook('ReporteVentas February 2023.xlsx')

sheet = workbook['Ventas']

last_row = max((r for r in range(1, sheet.max_row + 1) if sheet.cell(row=r, column=sheet.max_column).value), default=0)
if last_row:
    sheet.delete_rows(last_row, sheet.max_row - last_row + 1)

sheet.unmerge_cells('A1:H1')
sheet.unmerge_cells('A2:H2')
sheet.delete_rows(1, 6)
sheet.delete_rows(2)

skip_cols = ['C', 'E', 'F']

for col_idx in range(sheet.max_column, 0, -1):
    col_letter = get_column_letter(col_idx)
    if col_letter not in skip_cols:
        sheet.delete_cols(col_idx)

workbook.save('newModified.xlsx')

df = pd.read_excel('newModified.xlsx')

df_sorted = df.sort_values('Moneda', ascending=False)

df_sorted.to_excel('Modified Ventas Por Productos February 2023.xlsx', index=False)


import os

file_path = 'C:/Users/PCP/Desktop/Posto Python Scripts/newModified.xlsx'

os.remove(file_path)