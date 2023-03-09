import openpyxl


## This script is used to get the sum of each hour of operation. This data was gathered from
## The 'Ventas_Detallados' file/script

# -------------------- NOT USING THIS NOT USING THIS SCRIPT ------------------------------

# Load the workbook
workbook = openpyxl.load_workbook('Modified Ventas Detallados February 2023.xlsx')

# Select the worksheet
worksheet = workbook['Sheet1']

# Define the column names to process
number_col_name = "A"
total_col_name = "B"

# Get the column indices
number_col_index = openpyxl.utils.column_index_from_string(number_col_name)
total_col_index = openpyxl.utils.column_index_from_string(total_col_name)

# Initialize a dictionary to store the totals for each number
number_totals = {}

# Loop through the rows and add the totals for each number to the dictionary
for row in worksheet.iter_rows(min_row=2, values_only=True):
    number = row[number_col_index - 1]
    if number not in number_totals:
        number_totals[number] = 0
    number_totals[number] += row[total_col_index - 1]

# Create a list of unique numbers and their totals
unique_numbers = [(number, number_totals[number]) for number in number_totals]

# Sort the list by number
unique_numbers.sort(key=lambda x: x[0])

# Add a new column for the total number
total_number_col_name = "Hour"
total_number_col_index = openpyxl.utils.column_index_from_string("C")

# Write the headers to the worksheet
worksheet.cell(row=1, column=total_number_col_index, value=total_number_col_name)
worksheet.cell(row=1, column=total_number_col_index+1, value="Total")

# Write the unique numbers and their totals to the new column
for i, (number, total) in enumerate(unique_numbers):
    row_index = i + 2
    worksheet.cell(row=row_index, column=total_number_col_index, value=number)
    worksheet.cell(row=row_index, column=total_number_col_index+1, value=total)

# Save the workbook
workbook.save('Sum Of Each Hour February 2023.xlsx')




