import pandas as pd
from pandas import io
from pandas.io import excel
from pandas.io.formats import style

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from openpyxl import worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from openpyxl.utils.cell import  get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ----- THIS SCRIPT : To data cleanse all unused cells and columns of the excel sheet that provides the sum sold for each hour of operation.


workbook = openpyxl.load_workbook('Ventas Detallados November 2022.xlsx')

sheet = workbook['Costos']

# KEEP NOTE HERE TO CHANGE ROWS DEPENDING ON FILE
sheet.delete_rows(462, 478)

sheet.unmerge_cells('A1:L1')
sheet.unmerge_cells('A2:L2')
sheet.unmerge_cells('B4:D4')
sheet.unmerge_cells('F4:H4')
sheet.unmerge_cells('J4:L4')

sheet.unmerge_cells('B5:C5')
sheet.unmerge_cells('B6:C6')
sheet.unmerge_cells('B7:C7')
sheet.unmerge_cells('B8:C8')
sheet.unmerge_cells('B9:C9')
sheet.unmerge_cells('B10:C10')
sheet.unmerge_cells('B11:C11')
sheet.unmerge_cells('B12:C12')


sheet.unmerge_cells('F5:G5')
sheet.unmerge_cells('F6:G6')
sheet.unmerge_cells('F7:G7')
sheet.unmerge_cells('F8:G8')
sheet.unmerge_cells('F9:G9')
sheet.unmerge_cells('F10:G10')
sheet.unmerge_cells('F11:G11')
sheet.unmerge_cells('F12:G12')
sheet.unmerge_cells('F13:G13')
sheet.unmerge_cells('F14:G14')
sheet.unmerge_cells('F15:G15')


sheet.unmerge_cells('J5:K5')
sheet.unmerge_cells('J6:K6')
sheet.unmerge_cells('J7:K7')
sheet.unmerge_cells('J8:K8')
sheet.unmerge_cells('J9:K9')

sheet.unmerge_cells('A17:B17')

sheet.delete_rows(1,17)


skip_cols = ["B", "M"]  ## KEEP NOTE HERE. The column N for Total tends to change column placement depending on the file

for col_idx in range(sheet.max_column, 0, -1):
    col_letter = get_column_letter(col_idx)
    if col_letter not in skip_cols:
        sheet.delete_cols(col_idx)



# Find the index of the first empty cell in the last column
# last_column = sheet.max_column
# for row in range(1, sheet.max_row + 1):
#     if not sheet.cell(row=row, column=last_column).value:
#         break

# # Delete all the rows below the first empty cell
# if row < sheet.max_row:
#     sheet.delete_rows(row+1, sheet.max_row-row)


workbook.save('Modified Ventas Detallados November 2022.xlsx')

workbook = openpyxl.load_workbook('Modified Ventas Detallados November 2022.xlsx')

sheet = workbook['Costos']
# Loop through the cells in the column
for cell in sheet['A']:
    # Check if the cell contains a number
    if isinstance(cell.value, str) and cell.value.isnumeric():
        # Convert the cell value to a number
        cell.value = float(cell.value)

df = pd.read_excel('Modified Ventas Detallados November 2022.xlsx', sheet_name='Costos')

df_sorted = df.sort_values('Hora', ascending=True)

df_sorted.to_excel('Modified Ventas Detallados November 2022.xlsx', index=False)
# workbook.save('Modified Ventas Detallados November 2022.xlsx')

# # # This script is used to get the sum of each hour of operation. This data was gathered from the 'Ventas_Detallados' file/script


df = pd.read_excel('Modified Ventas Detallados November 2022.xlsx', sheet_name='Sheet1')

df_sorted = df.sort_values('Hora', ascending=True)

df_sorted = df.groupby('Hora')['Total'].sum().reset_index()

df_sorted = df_sorted.rename(columns={'Total': 'Total Number'})

df_sorted.to_excel('Sum Of Each Hour November 2022.xlsx', index=False)


# import os

# file_path = 'C:/Users/PCP/Desktop/Posto Python Scripts/NEW.xlsx'

# os.remove(file_path)