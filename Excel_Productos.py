import pandas as pd
from pandas import io
from pandas.io import excel
from pandas.io.formats import style

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from openpyxl import worksheet
from openpyxl.utils.dataframe import dataframe_to_rows


## This script will do data cleansing for the excel sheet that has products sold, its quantity and how many sold per hour of operation.

workbook = openpyxl.load_workbook('ventasProductos February 2022.xlsx')

sheet = workbook['Reporte de Productos']

sheet.unmerge_cells('A1:K1')
sheet.unmerge_cells('A2:K2')
sheet.delete_rows(1, 3)
# Delete a cell
# sheet['A1'] = None
# sheet['A2'] = None

skip_cols = ['A','C','D','Q', 'R', 'S','T','U','V', 'W', 'X','Y','Z','AA', 'AB']

for col_idx in range(sheet.max_column, 0, -1):  ## This will go through every column starting from the last one
    col_letter = get_column_letter(col_idx)     ## and checks every column letter
    if col_letter not in skip_cols:             ## if the column letter is not in the skip_cols variable that has the column letters to NOT skip
        sheet.delete_cols(col_idx)              ## Then it will delete those that are not in skip_cols

workbook.save('modified1.xlsx') ## Saving the file into a modified filed

df = pd.read_excel('modified1.xlsx') ## To sort the sheet based on a specified column I am using pandas dataframe

df_sorted = df.sort_values('Total', ascending=False) ## Sorting it by the 'Total' columns

df_sorted.to_excel('EXAMPLE Modified VentasProductos February 2023.xlsx', index=False)

import os

# Specify the path of the file you want to delete
file_path = 'C:/Users/PCP/Desktop/Posto Python Scripts/modified1.xlsx'

# Use the os module to delete the file
os.remove(file_path)


