from pandas import io
import openpyxl
from openpyxl import worksheet
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook('Client Transaction February 2023.xlsx')

sheet = workbook['Documento1']

sheet.unmerge_cells('A1:AH1')
sheet.unmerge_cells('A2:AH2')

sheet.delete_rows(1, 3)

skip_cols = ["H", "N", "W", "AA", "AB", "AD", "AK"]

for col_idx in range(sheet.max_column, 0, -1):
    col_letter = get_column_letter(col_idx)
    if col_letter not in skip_cols:
        sheet.delete_cols(col_idx)

names_to_drop = ['Consumidor Final']

for row in range(sheet.max_row, 1, -1):
    customer_name = sheet.cell(row=row, column=1).value
    discount_value = sheet.cell(row=row, column=6).value
    if customer_name in names_to_drop and discount_value <= 0:
        sheet.delete_rows(row)

# for row in range(sheet.max_row, 1, -1):
#     if sheet.cell(row=row, column=6).value <= 0:  # Assuming the discount column is in column C
#         sheet.delete_rows(row)

workbook.save('Modified Client Transaction February 2023.xlsx')


# need to edit so it cuts out all 'CONSUMIDOR FINAL' with a value <= 0 so we can see how much comp stuff we are giving.